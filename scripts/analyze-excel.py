#!/usr/bin/env python3
"""
Simple Excel analyzer to understand the file structure
"""

import openpyxl
import sys

def analyze_excel(file_path):
    try:
        print(f"Analyzing Excel file: {file_path}")
        wb = openpyxl.load_workbook(file_path, read_only=True)
        
        print(f"Worksheets: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"\nSheet: {sheet_name}")
            print(f"Dimensions: {sheet.max_row} rows, {sheet.max_column} columns")
            
            # Get headers
            headers = []
            for col in range(1, min(21, sheet.max_column + 1)):  # Max 20 columns
                cell_value = sheet.cell(row=1, column=col).value
                headers.append(str(cell_value) if cell_value else f"Col{col}")
            
            print(f"Headers: {headers}")
            
            # Sample data rows
            print("Sample data (first 3 rows):")
            for row_num in range(2, min(5, sheet.max_row + 1)):
                row_data = []
                for col in range(1, min(11, sheet.max_column + 1)):  # First 10 columns
                    cell_value = sheet.cell(row=row_num, column=col).value
                    row_data.append(str(cell_value) if cell_value else "")
                print(f"  Row {row_num}: {row_data}")
            
            if sheet.max_row > 4:
                print(f"... and {sheet.max_row - 4} more rows")
        
        wb.close()
        
    except Exception as e:
        print(f"Error analyzing Excel file: {e}")

if __name__ == "__main__":
    file_path = sys.argv[1] if len(sys.argv) > 1 else "JANS STOCK 18-Jul-25x505 USD_1753036569196.XLSX"
    analyze_excel(file_path)