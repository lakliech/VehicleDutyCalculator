from openpyxl import load_workbook
import sys

try:
    # Read the Excel file
    file_path = "attached_assets/product and features_1752714449672.xlsx"
    
    # Load workbook
    wb = load_workbook(file_path)
    print("Sheet names:", wb.sheetnames)
    
    # Read each sheet
    for sheet_name in wb.sheetnames:
        print(f"\n=== Sheet: {sheet_name} ===")
        ws = wb[sheet_name]
        
        # Get dimensions
        max_row = ws.max_row
        max_col = ws.max_column
        print(f"Dimensions: {max_row} rows, {max_col} columns")
        
        # Read headers
        headers = []
        for col in range(1, max_col + 1):
            cell_value = ws.cell(row=1, column=col).value
            headers.append(cell_value)
        print("Headers:", headers)
        
        # Read all rows of data
        print("All data:")
        for row in range(1, max_row + 1):
            row_data = []
            for col in range(1, max_col + 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(cell_value)
            if any(row_data):  # Only print non-empty rows
                print(f"Row {row}: {row_data}")
        
        print("\n" + "="*50)

except Exception as e:
    print(f"Error reading Excel file: {e}")